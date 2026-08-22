# Copyright (c) 2026, MBWD and contributors
# For license information, please see license.txt

"""Đọc file Excel BOM Template của HKLED thành đặc tả JSON (PM-TASK-00110).

Nguồn: Google Sheet trong mô tả PM-TASK-00110, 6 sheet = 6 mặt hàng cha.
Chạy lại khi khách sửa file:

    python -m mbwnext_hkled.data.doc_bom_sheet <file.xlsx> > .../bom_template/spec.json

## Bố cục file

Mỗi sheet gồm bảng thành phần rồi tới các khối rule. Hai dạng khối:

    <Thành phần> | Mọi biến thể đều chọn | <Mã NVL>          ← một dòng, không điều kiện

    <Thành phần> | Đặc tính thành phẩm |   |   | Mã NVL      ← dòng mở khối
                 | <đt 1> | <đt 2> | …                       ← tên các đặc tính
                 | <gt>   | <gt>   | … | <Mã NVL>            ← từng dòng rule

## Quy ước ô — chỗ dễ đọc sai nhất

Khách điền **đủ mọi ô điều kiện trên mọi dòng**. Cả file chỉ có hai chỗ lệch, và hai
chỗ đó nghĩa khác hẳn nhau, nên không có luật chung nào đúng cho cả hai:

1. `Còn lại` (DP01S/Module) — ghi ở dòng đầu của khối cuối, các dòng sau bỏ trống.
   Nghĩa: *mọi tổ hợp mà các khối phía trên chưa bắt*. Toàn khối mang nghĩa này.
2. Ô trống rải rác (DP01S/Cầu đấu, cột `Nguồn` chỉ điền 1/56 dòng) — **không suy được
   nghĩa**. Hàm này `raise`, không đoán.

Vì vậy: gặp cột điều kiện điền dở mà không phải dạng (1) thì dừng và báo lỗi. Đoán ở đây
là đoán ra một BOM sai, mà BOM sai chảy thẳng vào lệnh sản xuất và phiếu xuất kho.

## Đổi tên đặc tính giữa file và phần mềm

Sheet và danh mục đặt tên khác nhau ở vài chỗ; ánh xạ do Thắng chốt 18/08 trên
PM-TASK-00110 (`DOI_TEN_DAC_TINH`). Đổi tên thành phần cũng vậy (`DOI_TEN_TP`).
"""

import json
import re
import sys

MO_KHOI = "Đặc tính thành phẩm"
MOI_BIEN_THE = "Mọi biến thể đều chọn"
CON_LAI = "Còn lại"
TAT_CA = "tất cả"
KHONG_DUNG = "Không sử dụng"
COT_NVL = "Mã NVL"

# Khách viết gộp nhiều giá trị vào một ô: "HKLED Dim 1 Cấp và HKLED Dim 5 Cấp".
NOI_GIA_TRI = " và "
# Cách viết khác của "Còn lại", dùng ở cột Nguồn của khối Cầu đấu (DP01S).
# ⚠ So sánh KHÔNG phân biệt hoa thường: khách viết cả "tất cả" lẫn "Tất cả" trong cùng
# một sheet. Khớp cứng theo chữ thì rule biến thành điều kiện "Công suất = Tất cả" —
# không biến thể nào có giá trị đó nên rule chết lặng, không báo gì.
CON_LAI_KHAC = ("các loại còn lại",)


def _la(gia_tri, *mau):
	return gia_tri.strip().lower() in [m.strip().lower() for m in mau]

# Chốt của Thắng 18/08: tên trong sheet -> tên bản ghi BOM Component đã có trên phần mềm.
DOI_TEN_TP = {
	"Gioăng": "Gioăng chip",
	"Ốc vít cố định Module": "Ốc vít cố định module",
}

# Chốt của Thắng 18/08: sheet gọi một tên, biến thể trên phần mềm mang tên khác nhưng
# CÙNG bộ giá trị. Trên phần mềm đây là hai Item Attribute riêng biệt nên phải ánh xạ
# tường minh, không tự suy theo giá trị giống nhau.
DOI_TEN_DAC_TINH = {
	"DP01S": {"Loại LED": "Chip LED"},
	"VDP0X": {"Kiểu lắp": "Phân loại vỏ"},
}

# Sheet ghi kèm đơn vị, phần mềm lưu số trần (chữ W/LED nằm ở phần viết tắt của đặc tính).
DON_VI = {"Công suất": "W", "Số lượng LED": "LED"}


def chuan_gia_tri(dac_tinh, gia_tri):
	don_vi = DON_VI.get(dac_tinh)
	if don_vi:
		khop = re.fullmatch(r"(\d+(?:\.\d+)?)\s*" + don_vi, gia_tri)
		if khop:
			return khop.group(1)
	return gia_tri


def _o(row, i):
	return "" if i >= len(row) or row[i] is None else str(row[i]).strip()


def _doc_sheet(ten_sheet, rows):
	doi_dt = DOI_TEN_DAC_TINH.get(ten_sheet, {})
	item_cha = ""
	thanh_phan = []
	rules = []
	i = 0
	trong_bang_tp = False
	tp_gon = ""   # thành phần của khối rule một dòng đang mở

	while i < len(rows):
		row = rows[i]
		c0, c1 = _o(row, 0), _o(row, 1)

		if c0 == "Item cha":
			item_cha = _o(row, 1)
			i += 1
			continue

		if c1 == "Thành phần BOM":  # tiêu đề bảng thành phần
			trong_bang_tp = True
			i += 1
			continue

		# Rule một dòng: <Thành phần> | <Đặc tính>: <Giá trị> | <Mã NVL>
		# Dòng kế tiếp cùng thành phần bỏ trống cột đầu, nên phải nhớ thành phần đang mở.
		khop_gon = re.fullmatch(r"([^:]{1,40}):\s*(.+)", c1) if _o(row, 2) else None
		if khop_gon and (c0 or tp_gon):
			if c0:
				tp_gon = DOI_TEN_TP.get(c0, c0)
			dt_goc = khop_gon.group(1).strip()
			ten_dt = doi_dt.get(dt_goc, dt_goc)
			rules.append({
				"thanh_phan": tp_gon, "moi_bien_the": False,
				"cond": {ten_dt: chuan_gia_tri(dt_goc, khop_gon.group(2).strip())},
				"nvl": _o(row, 2),
			})
			trong_bang_tp = False
			i += 1
			continue
		if c0 or c1 == MO_KHOI or c1 == MOI_BIEN_THE:
			tp_gon = ""

		if c1 == MOI_BIEN_THE:
			rules.append(
				{"thanh_phan": DOI_TEN_TP.get(c0, c0), "moi_bien_the": True, "cond": {}, "nvl": _o(row, 2)}
			)
			trong_bang_tp = False
			i += 1
			continue

		if c1 == MO_KHOI:
			tp = DOI_TEN_TP.get(c0, c0)
			cot_nvl = next((k for k in range(len(row)) if _o(row, k) == COT_NVL), None)
			if cot_nvl is None:
				raise ValueError(f"{ten_sheet}: khối {tp!r} (dòng {i + 1}) không có cột {COT_NVL!r}")
			hdr = rows[i + 1]
			dac_tinh = [_o(hdr, k) for k in range(1, cot_nvl) if _o(hdr, k)]
			i += 2
			khoi = []
			while i < len(rows) and not _o(rows[i], 0) and _o(rows[i], 1) != MO_KHOI:
				r = rows[i]
				if not any(_o(r, k) for k in range(len(r))):
					break
				khoi.append([_o(r, 1 + k) for k in range(len(dac_tinh))] + [_o(r, cot_nvl)])
				i += 1
			rules.extend(_doc_khoi(ten_sheet, tp, dac_tinh, khoi, doi_dt))
			trong_bang_tp = False
			continue

		if trong_bang_tp and c1:
			kieu = {"cố định": "Cố Định", "theo rule": "Theo Rule",
				"số lượng theo công thức": "Số Lượng Theo Công Thức"}.get(_o(row, 2).lower())
			if kieu:
				# Ô SL nằm sau nhãn "SL" và có thể trải ra NHIỀU cột khi số lượng đổi theo
				# đặc tính: "Kiểu đấu: Cầu đấu SL 1" | "Kiểu đấu: Dây diện SL 30". Giữ
				# nguyên văn cả cụm — cắt bớt là mất mất một nhánh số lượng.
				sl = []
				if _o(row, 4) == "SL":
					sl = [_o(row, k) for k in range(5, len(row)) if _o(row, k)]
				thanh_phan.append({
					"thanh_phan": DOI_TEN_TP.get(c1, c1),
					"kieu": kieu,
					"nvl": _o(row, 3),
					"sl_tho": " | ".join(sl),
				})
		i += 1

	return {"item_cha": item_cha or ten_sheet, "thanh_phan": thanh_phan, "rules": rules}


def _doc_khoi(ten_sheet, tp, dac_tinh, khoi, doi_dt):
	"""Một khối rule -> danh sách rule. Kiểm quy ước ô ngay tại đây."""
	so_dong = len(khoi)
	rules = []
	con_lai_tu = {}   # cột -> dòng bắt đầu phần 'Còn lại'

	bo_cot = set()
	for k, dt in enumerate(dac_tinh):
		co = sum(1 for r in khoi if r[k])
		if co == so_dong:
			continue
		if co == 0:
			# Cột có tiêu đề nhưng không dòng nào điền -> đặc tính không được dùng ở khối này.
			bo_cot.add(k)
			continue
		# Dạng (1): 'Còn lại' ghi đúng một lần, mọi dòng TRƯỚC nó đều có giá trị và mọi
		# dòng SAU nó đều trống -> cả phần đuôi thuộc về 'Còn lại'.
		vt = [j for j, r in enumerate(khoi) if _la(r[k], CON_LAI, *CON_LAI_KHAC)]
		if len(vt) == 1:
			p = vt[0]
			if all(khoi[j][k] for j in range(p)) and not any(khoi[j][k] for j in range(p + 1, so_dong)):
				con_lai_tu[k] = p
				continue
		raise ValueError(
			f"{ten_sheet}: khối {tp!r}, cột điều kiện {dt!r} chỉ điền {co}/{so_dong} dòng"
			f" và không phải dạng 'Còn lại'. Không suy được ý — cần khách điền đủ hoặc"
			f" xác nhận cách hiểu trước khi dựng rule."
		)

	for j, r in enumerate(khoi):
		cond = {}
		for k, dt in enumerate(dac_tinh):
			ten = doi_dt.get(dt, dt)
			if k in bo_cot:
				continue
			if k in con_lai_tu and j >= con_lai_tu[k]:
				cond[ten] = CON_LAI
				continue
			gt = r[k]
			if _la(gt, TAT_CA):  # không ràng buộc đặc tính này
				continue
			if _la(gt, CON_LAI, *CON_LAI_KHAC):
				cond[ten] = CON_LAI
				continue
			if NOI_GIA_TRI in gt:  # nhiều giá trị trong một ô -> OR
				cond[ten] = [chuan_gia_tri(dt, x.strip()) for x in gt.split(NOI_GIA_TRI) if x.strip()]
				continue
			cond[ten] = chuan_gia_tri(dt, gt)
		rules.append({"thanh_phan": tp, "moi_bien_the": False, "cond": cond, "nvl": r[-1]})
	return rules


def doc(path):
	"""Trả `{"sheets": {...}, "loi": {ten_sheet: thông báo}}`.

	Lỗi gom theo sheet chứ không dừng cả file: một sheet khách điền dở không nên chặn
	năm sheet còn lại. Sheet nào có lỗi thì KHÔNG nằm trong `sheets` — không dựng nửa vời.
	"""
	import openpyxl

	wb = openpyxl.load_workbook(path)
	sheets, loi = {}, {}
	for s in wb.sheetnames:
		try:
			sheets[s] = _doc_sheet(s, list(wb[s].iter_rows(values_only=True)))
		except ValueError as e:
			loi[s] = str(e)
	return {"sheets": sheets, "loi": loi}


if __name__ == "__main__":
	ra = doc(sys.argv[1])
	for s, m in ra["loi"].items():
		print(f"⚠ BỎ QUA sheet {s}: {m}", file=sys.stderr)
	print(json.dumps(ra, ensure_ascii=False, indent=1))
